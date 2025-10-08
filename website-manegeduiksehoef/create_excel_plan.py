import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# Data voor het stappenplan
data = {
    'Stap': [
        '1', '1.1', '1.2', '1.3', '1.4', '1.5',
        '2', '2.1', '2.2', '2.3', '2.4', '2.5',
        '3', '3.1', '3.2', '3.3', '3.4',
        '4', '4.1', '4.2', '4.3',
        '5', '5.1', '5.2', '5.3'
    ],
    'Taak': [
        'Database opzetten',
        'Supabase account aanmaken',
        'Nieuwe project aanmaken',
        'Database tabellen maken',
        'Product data migreren',
        'Database connectie testen',
        'Stripe betaling',
        'Stripe account aanmaken',
        'Stripe keys ophalen',
        'Stripe package installeren',
        'Checkout integratie',
        'Test betalingen',
        'Vercel deployment',
        'Vercel account aanmaken',
        'GitHub repository koppelen',
        'Environment variables instellen',
        'Eerste deployment',
        'Email notificaties',
        'Resend account aanmaken',
        'Email templates maken',
        'Email integratie',
        'SEO & Analytics',
        'Meta tags toevoegen',
        'Google Analytics',
        'Sitemap genereren'
    ],
    'Status': [
        'NOG TE DOEN', 'NOG TE DOEN', 'NOG TE DOEN', 'NOG TE DOEN', 'NOG TE DOEN', 'NOG TE DOEN',
        'NOG TE DOEN', 'NOG TE DOEN', 'NOG TE DOEN', 'NOG TE DOEN', 'NOG TE DOEN', 'NOG TE DOEN',
        'NOG TE DOEN', 'NOG TE DOEN', 'NOG TE DOEN', 'NOG TE DOEN', 'NOG TE DOEN',
        'NOG TE DOEN', 'NOG TE DOEN', 'NOG TE DOEN', 'NOG TE DOEN',
        'NOG TE DOEN', 'NOG TE DOEN', 'NOG TE DOEN', 'NOG TE DOEN'
    ],
    'Tijd': [
        '45 min', '5 min', '5 min', '15 min', '15 min', '5 min',
        '60 min', '10 min', '5 min', '5 min', '30 min', '10 min',
        '30 min', '5 min', '5 min', '10 min', '10 min',
        '30 min', '5 min', '15 min', '10 min',
        '35 min', '15 min', '10 min', '10 min'
    ],
    'Prioriteit': [
        'Kritiek', 'Kritiek', 'Kritiek', 'Kritiek', 'Kritiek', 'Kritiek',
        'Kritiek', 'Kritiek', 'Kritiek', 'Kritiek', 'Kritiek', 'Kritiek',
        'Kritiek', 'Kritiek', 'Kritiek', 'Kritiek', 'Kritiek',
        'Hoog', 'Hoog', 'Hoog', 'Hoog',
        'Gemiddeld', 'Gemiddeld', 'Gemiddeld', 'Gemiddeld'
    ],
    'Notities': [
        'Supabase database',
        'https://supabase.com',
        'Project naam: mee-hestar-webshop',
        'Products, Orders, Customers',
        'Van mock data naar database',
        'Verificatie',
        'Betalingsintegratie',
        'https://stripe.com',
        'Publishable & Secret key',
        'npm install stripe',
        'API routes maken',
        'Verificatie',
        'Live hosting',
        'https://vercel.com',
        'Code uploaden',
        'Database & Stripe keys',
        'Live zetten',
        'Order bevestigingen',
        'https://resend.com',
        'Order bevestiging',
        'API routes',
        'Optimalisatie',
        'SEO optimalisatie',
        'Tracking',
        'SEO'
    ]
}

# Maak DataFrame
df = pd.DataFrame(data)

# Maak Excel bestand
wb = Workbook()
ws = wb.active
ws.title = "Stappenplan Webshop"

# Voeg titel toe
ws['A1'] = "📋 STAPPENPLAN - Mee Hestar Webshop Live"
ws['A1'].font = Font(size=16, bold=True, color="2D6A4F")
ws.merge_cells('A1:F1')

# Voeg subtitel toe
ws['A2'] = f"🎯 Doel: Webshop live in 3 uur | 📅 Gemaakt op: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
ws['A2'].font = Font(size=12, color="666666")
ws.merge_cells('A2:F2')

# Voeg lege rij toe
ws['A3'] = ""

# Voeg headers toe
headers = ['Stap', 'Taak', 'Status', 'Tijd', 'Prioriteit', 'Notities']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="2D6A4F", end_color="2D6A4F", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Voeg data toe
for r in dataframe_to_rows(df, index=False, header=False):
    ws.append(r)

# Styling voor de data
for row in range(5, len(df) + 5):
    # Status kolom styling
    status_cell = ws.cell(row=row, column=3)
    if status_cell.value == 'NOG TE DOEN':
        status_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        status_cell.font = Font(color="856404")
    elif status_cell.value == 'AFGEVINKT':
        status_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        status_cell.font = Font(color="155724")
    
    # Prioriteit kolom styling
    priority_cell = ws.cell(row=row, column=5)
    if priority_cell.value == 'Kritiek':
        priority_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        priority_cell.font = Font(color="721C24")
    elif priority_cell.value == 'Hoog':
        priority_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        priority_cell.font = Font(color="856404")
    elif priority_cell.value == 'Gemiddeld':
        priority_cell.fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
        priority_cell.font = Font(color="0C5460")

# Voeg voortgang sectie toe
start_row = len(df) + 7
ws.cell(row=start_row, column=1, value="📊 VOORTGANG").font = Font(bold=True, size=14, color="2D6A4F")
ws.cell(row=start_row + 1, column=1, value="⏳ NOG TE DOEN:").font = Font(bold=True)
ws.cell(row=start_row + 1, column=2, value=len(df)).font = Font(bold=True, color="856404")
ws.cell(row=start_row + 2, column=1, value="✅ AFGEVINKT:").font = Font(bold=True)
ws.cell(row=start_row + 2, column=2, value=0).font = Font(bold=True, color="155724")
ws.cell(row=start_row + 3, column=1, value="TOTAAL:").font = Font(bold=True)
ws.cell(row=start_row + 3, column=2, value=len(df)).font = Font(bold=True, color="2D6A4F")

# Voeg tijdsplan toe
start_row += 5
ws.cell(row=start_row, column=1, value="🎯 TIJDSPLAN").font = Font(bold=True, size=14, color="2D6A4F")

tijdsplan_data = [
    ['Database', '45 min', '⏳'],
    ['Betaling', '60 min', '⏳'],
    ['Deployment', '30 min', '⏳'],
    ['Email', '30 min', '⏳'],
    ['SEO', '35 min', '⏳'],
    ['TOTAAL', '3 uur 20 min', '⏳']
]

for i, (fase, tijd, status) in enumerate(tijdsplan_data):
    ws.cell(row=start_row + 1 + i, column=1, value=fase).font = Font(bold=True)
    ws.cell(row=start_row + 1 + i, column=2, value=tijd)
    ws.cell(row=start_row + 1 + i, column=3, value=status).font = Font(color="856404")

# Voeg handige links toe
start_row += 8
ws.cell(row=start_row, column=1, value="🔗 HANDIGE LINKS").font = Font(bold=True, size=14, color="2D6A4F")

links_data = [
    ['Supabase', 'https://supabase.com', 'Database'],
    ['Stripe', 'https://stripe.com', 'Betaling'],
    ['Vercel', 'https://vercel.com', 'Hosting'],
    ['Resend', 'https://resend.com', 'Email'],
    ['Google Analytics', 'https://analytics.google.com', 'Analytics']
]

for i, (naam, url, beschrijving) in enumerate(links_data):
    ws.cell(row=start_row + 1 + i, column=1, value=naam).font = Font(bold=True)
    ws.cell(row=start_row + 1 + i, column=2, value=url).font = Font(color="0066CC")
    ws.cell(row=start_row + 1 + i, column=3, value=beschrijving)

# Auto-adjust kolommen
for column in ws.columns:
    max_length = 0
    column_letter = None
    for cell in column:
        if hasattr(cell, 'column_letter'):
            column_letter = cell.column_letter
            break
    if column_letter:
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

# Voeg borders toe
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in ws.iter_rows(min_row=4, max_row=len(df) + 4, min_col=1, max_col=6):
    for cell in row:
        cell.border = thin_border

# Sla bestand op
filename = f"Stappenplan_Mee_Hestar_Webshop_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
wb.save(filename)

print(f"✅ Excel bestand aangemaakt: {filename}")
print(f"📁 Bestandspad: {filename}")
print(f"📊 Bevat {len(df)} taken verdeeld over 5 fases")
print(f"⏱️ Geschatte tijd: 3 uur 20 minuten")
